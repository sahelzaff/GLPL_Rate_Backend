from flask import Blueprint, request, jsonify, send_file, make_response
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from middleware.auth import admin_required
from config.database import Database
from datetime import datetime
from flask_cors import cross_origin

bulk_upload_routes = Blueprint('bulk_upload_routes', __name__)
db = Database.get_instance().db

def create_ports_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Ports"
    
    # Add headers exactly matching database schema
    headers = [
        'port_code',    # Primary identifier
        'port_name',    # Name of the port
        'country',      # Country where port is located
        'region',       # Region/State within country
        'created_at',   # Creation timestamp (will be auto-filled)
        'updated_at'    # Last update timestamp (will be auto-filled)
    ]
    
    # Add headers
    ws.append(headers)
    
    # Get existing ports from database
    existing_ports = list(db.ports.find({}, {'_id': 0}))  # Exclude _id field
    
    # Add existing ports data
    for port in existing_ports:
        ws.append([
            port.get('port_code', ''),
            port.get('port_name', ''),
            port.get('country', ''),
            port.get('region', ''),
            str(port.get('created_at', '')),
            str(port.get('updated_at', ''))
        ])
    
    # Add note about template usage
    ws.append([])  # Empty row
    note = ['Note: Add new ports below the existing data. Do not modify existing ports data.']
    ws.append(note)
    
    # Style the template
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C6082C', end_color='C6082C', fill_type='solid')
    note_font = Font(bold=True, color='FF0000')
    existing_data_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    
    # Style headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Style existing data
    for row in range(2, len(existing_ports) + 2):
        for cell in ws[row]:
            cell.fill = existing_data_fill
    
    # Style note
    note_row = len(existing_ports) + 3
    ws[f'A{note_row}'].font = note_font
    ws.merge_cells(f'A{note_row}:F{note_row}')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def create_shipping_lines_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipping Lines"
    
    # Add headers exactly matching database schema
    headers = [
        'name',              # Company name
        'contact_email',     # Primary contact
        'website',          # Company website
        'created_at',       # Creation timestamp
        'updated_at'        # Last update timestamp
    ]
    
    # Add headers
    ws.append(headers)
    
    # Get existing shipping lines from database
    existing_lines = list(db.shipping_lines.find({}, {'_id': 0}))
    
    # Add existing shipping lines data
    for line in existing_lines:
        ws.append([
            line.get('name', ''),
            line.get('contact_email', ''),
            line.get('website', ''),
            str(line.get('created_at', '')),
            str(line.get('updated_at', ''))
        ])
    
    # Add note about template usage
    ws.append([])  # Empty row
    note = ['Note: Add new shipping lines below the existing data. Do not modify existing data.']
    ws.append(note)
    
    # Style the template
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C6082C', end_color='C6082C', fill_type='solid')
    note_font = Font(bold=True, color='FF0000')
    existing_data_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    
    # Apply styles
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Style existing data
    for row in range(2, len(existing_lines) + 2):
        for cell in ws[row]:
            cell.fill = existing_data_fill
    
    # Style note
    note_row = len(existing_lines) + 3
    ws[f'A{note_row}'].font = note_font
    ws.merge_cells(f'A{note_row}:E{note_row}')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def create_rates_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"
    
    # Add headers exactly matching database schema
    headers = [
        'shipping_line_id',  # Shipping line reference
        'pol_id',           # Port of loading
        'pod_id',           # Port of discharge
        'container_type',   # Type of container
        'rate',            # Rate amount
        'valid_from',      # Validity start date
        'valid_to',        # Validity end date
    ]
    
    # Add headers
    ws.append(headers)
    
    # Get existing rates from database with populated references
    pipeline = [
        {
            '$lookup': {
                'from': 'shipping_lines',
                'localField': 'shipping_line_id',
                'foreignField': '_id',
                'as': 'shipping_line'
            }
        },
        {
            '$lookup': {
                'from': 'ports',
                'localField': 'pol_id',
                'foreignField': '_id',
                'as': 'pol'
            }
        },
        {
            '$lookup': {
                'from': 'ports',
                'localField': 'pod_id',
                'foreignField': '_id',
                'as': 'pod'
            }
        },
        {
            '$unwind': '$shipping_line'
        },
        {
            '$unwind': '$pol'
        },
        {
            '$unwind': '$pod'
        }
    ]
    
    existing_rates = list(db.rates.aggregate(pipeline))
    
    # Add existing rates data
    for rate in existing_rates:
        ws.append([
            str(rate['shipping_line_id']),
            str(rate['pol_id']),
            str(rate['pod_id']),
            rate.get('container_type', ''),
            rate.get('rate', ''),
            rate.get('valid_from', '').strftime('%Y-%m-%d') if isinstance(rate.get('valid_from'), datetime) else str(rate.get('valid_from', '')),
            rate.get('valid_to', '').strftime('%Y-%m-%d') if isinstance(rate.get('valid_to'), datetime) else str(rate.get('valid_to', ''))
        ])
    
    # Add helper information
    ws.append([])  # Empty row
    ws.append(['Available Shipping Lines:'])
    shipping_lines = list(db.shipping_lines.find({}, {'_id': 1, 'name': 1}))
    for sl in shipping_lines:
        ws.append([f"ID: {str(sl['_id'])} - {sl['name']}"])
    
    ws.append([])
    ws.append(['Available Ports:'])
    ports = list(db.ports.find({}, {'_id': 1, 'port_code': 1, 'port_name': 1}))
    for port in ports:
        ws.append([f"ID: {str(port['_id'])} - {port['port_code']} ({port['port_name']})"])
    
    ws.append([])
    ws.append(['Container Types:'])
    container_types = ['20GP', '40GP', '40HC', '20RF', '40RF']  # Add your container types
    for ct in container_types:
        ws.append([ct])
    
    ws.append([])
    note = ['Note: Add new rates below the existing data. Format dates as YYYY-MM-DD. Do not modify existing data.']
    ws.append(note)
    
    # Style the template
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C6082C', end_color='C6082C', fill_type='solid')
    note_font = Font(bold=True, color='FF0000')
    existing_data_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    helper_font = Font(bold=True, color='0000FF')
    
    # Apply styles
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Style existing data
    for row in range(2, len(existing_rates) + 2):
        for cell in ws[row]:
            cell.fill = existing_data_fill
    
    # Style helper information
    current_row = len(existing_rates) + 3
    ws[f'A{current_row}'].font = helper_font  # Shipping Lines header
    current_row += len(shipping_lines) + 2
    ws[f'A{current_row}'].font = helper_font  # Ports header
    current_row += len(ports) + 2
    ws[f'A{current_row}'].font = helper_font  # Container Types header
    
    # Style note
    note_row = current_row + len(container_types) + 2
    ws[f'A{note_row}'].font = note_font
    ws.merge_cells(f'A{note_row}:G{note_row}')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@bulk_upload_routes.route('/api/templates/<type>', methods=['GET', 'OPTIONS'])
def download_template(type):
    try:
        if request.method == 'OPTIONS':
            return make_response('', 200)

        if type == 'ports':
            buffer = create_ports_template()
            filename = 'ports_template.xlsx'
        elif type == 'shipping-lines':
            buffer = create_shipping_lines_template()
            filename = 'shipping_lines_template.xlsx'
        elif type == 'rates':
            buffer = create_rates_template()
            filename = 'rates_template.xlsx'
        else:
            return jsonify({"error": "Invalid template type"}), 400

        # Send file without adding extra CORS headers
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Download error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@bulk_upload_routes.route('/api/<type>/preview', methods=['POST', 'OPTIONS'])
@cross_origin()
@admin_required
def preview_upload(type):
    if request.method == 'OPTIONS':
        return make_response('', 200)
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files['file']
        if not file.filename.endswith('.xlsx'):
            return jsonify({"error": "Invalid file format"}), 400

        # Read Excel file
        df = pd.read_excel(file)
        
        if type == 'ports':
            required_columns = ['port_code', 'port_name', 'country', 'region']
            display_columns = required_columns[:4]  # Exclude created_at and updated_at
            
            # Validate required columns
            missing_columns = [col for col in required_columns[:3] if col not in df.columns]
            if missing_columns:
                return jsonify({"error": f"Missing required columns: {', '.join(missing_columns)}"}), 400
            
            # Get existing port codes
            existing_port_codes = {port['port_code'] for port in db.ports.find({}, {'port_code': 1})}
            
            # Clean data and filter new ports
            preview_data = []
            for index, row in df.iterrows():
                if pd.isna(row['port_code']) or pd.isna(row['port_name']) or pd.isna(row['country']):
                    continue
                
                port_code = str(row['port_code']).strip().upper()
                
                # Only include ports that don't exist in database
                if port_code not in existing_port_codes:
                    preview_data.append({
                        'port_code': port_code,
                        'port_name': str(row['port_name']).strip(),
                        'country': str(row['country']).strip(),
                        'region': str(row.get('region', '')).strip() if not pd.isna(row.get('region')) else ''
                    })
            
            return jsonify({
                "columns": display_columns,
                "data": preview_data,
                "total_records": len(preview_data)
            })
            
        elif type == 'shipping-lines':
            required_columns = ['name', 'contact_email', 'website']
            display_columns = required_columns  # Exclude created_at and updated_at
            
            # Validate required columns
            missing_columns = [col for col in required_columns[:2] if col not in df.columns]
            if missing_columns:
                return jsonify({"error": f"Missing required columns: {', '.join(missing_columns)}"}), 400
            
            # Get existing shipping line names
            existing_names = {line['name'] for line in db.shipping_lines.find({}, {'name': 1})}
            
            # Clean data and filter new shipping lines
            preview_data = []
            for index, row in df.iterrows():
                if pd.isna(row['name']) or pd.isna(row['contact_email']):
                    continue
                
                name = str(row['name']).strip()
                
                # Only include shipping lines that don't exist in database
                if name not in existing_names:
                    preview_data.append({
                        'name': name,
                        'contact_email': str(row['contact_email']).strip(),
                        'website': str(row.get('website', '')).strip() if not pd.isna(row.get('website')) else ''
                    })
            
            return jsonify({
                "columns": display_columns,
                "data": preview_data,
                "total_records": len(preview_data)
            })
            
        elif type == 'rates':
            required_columns = ['shipping_line_id', 'pol_id', 'pod_id', 'container_type', 'rate', 'valid_from', 'valid_to']
            display_columns = required_columns + ['status', 'current_rate']  # Add status column to show if new/updated
            
            # Validate required columns
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({"error": f"Missing required columns: {', '.join(missing_columns)}"}), 400
            
            # Get existing rates with their current values
            existing_rates = {}
            for rate in db.rates.find({}, {
                'shipping_line_id': 1, 'pol_id': 1, 'pod_id': 1, 
                'container_type': 1, 'rate': 1, 'valid_from': 1, 'valid_to': 1
            }):
                key = f"{rate['shipping_line_id']}-{rate['pol_id']}-{rate['pod_id']}-{rate['container_type']}"
                if key not in existing_rates or rate['valid_to'] > existing_rates[key]['valid_to']:
                    existing_rates[key] = rate
            
            # Clean data and identify changes
            preview_data = []
            for index, row in df.iterrows():
                if any(pd.isna(row[col]) for col in required_columns):
                    continue
                
                # Clean and prepare the rate data
                rate_data = {
                    'shipping_line_id': str(row['shipping_line_id']).strip(),
                    'pol_id': str(row['pol_id']).strip(),
                    'pod_id': str(row['pod_id']).strip(),
                    'container_type': str(row['container_type']).strip(),
                    'rate': float(row['rate']),
                    'valid_from': pd.to_datetime(row['valid_from']).strftime('%Y-%m-%d'),
                    'valid_to': pd.to_datetime(row['valid_to']).strftime('%Y-%m-%d')
                }
                
                # Generate key for comparison
                rate_key = f"{rate_data['shipping_line_id']}-{rate_data['pol_id']}-{rate_data['pod_id']}-{rate_data['container_type']}"
                
                # Check if this is a new route or rate update
                if rate_key in existing_rates:
                    existing_rate = existing_rates[rate_key]
                    if existing_rate['rate'] != rate_data['rate']:
                        preview_data.append({
                            **rate_data,
                            'status': 'Update',
                            'current_rate': existing_rate['rate']
                        })
                else:
                    preview_data.append({
                        **rate_data,
                        'status': 'New',
                        'current_rate': '-'
                    })
            
            return jsonify({
                "columns": display_columns,
                "data": preview_data,
                "total_records": len(preview_data)
            })
        
        else:
            return jsonify({"error": "Invalid upload type"}), 400

    except Exception as e:
        print(f"Preview error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@bulk_upload_routes.route('/api/<type>/bulk', methods=['POST', 'OPTIONS'])
@cross_origin()
@admin_required
def bulk_upload(type):
    if request.method == 'OPTIONS':
        return make_response('', 200)
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files['file']
        df = pd.read_excel(file)

        if type == 'ports':
            result = bulk_upload_ports(df)
        elif type == 'shipping-lines':
            result = bulk_upload_shipping_lines(df)
        elif type == 'rates':
            result = bulk_upload_rates(df)
        else:
            return jsonify({"error": "Invalid upload type"}), 400

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

def bulk_upload_ports(df):
    required_columns = ['port_code', 'port_name', 'country']
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")

    # Get existing port codes
    existing_port_codes = {port['port_code'] for port in db.ports.find({}, {'port_code': 1})}
    
    records = df.to_dict('records')
    inserted = 0
    updated = 0
    errors = []

    for index, record in enumerate(records):
        try:
            # Skip empty rows
            if pd.isna(record['port_code']) or pd.isna(record['port_name']) or pd.isna(record['country']):
                continue
                
            # Clean data
            cleaned_record = {
                'port_code': str(record['port_code']).strip().upper(),
                'port_name': str(record['port_name']).strip(),
                'country': str(record['country']).strip(),
                'region': str(record.get('region', '')).strip() if not pd.isna(record.get('region')) else '',
                'updated_at': datetime.utcnow()
            }
            
            # Only process new ports
            if cleaned_record['port_code'] not in existing_port_codes:
                cleaned_record['created_at'] = datetime.utcnow()
                
                result = db.ports.insert_one(cleaned_record)
                if result.inserted_id:
                    inserted += 1
                    
        except Exception as e:
            errors.append(f"Row {index + 1}: {str(e)}")
            print(f"Error processing record: {record}")
            print(f"Error: {str(e)}")

    return {
        "message": "Bulk upload completed",
        "inserted": inserted,
        "skipped": len(existing_port_codes),
        "errors": errors if errors else None
    }

def bulk_upload_shipping_lines(df):
    required_columns = ['name', 'contact_email']
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")

    # Get existing shipping line names
    existing_names = {line['name'] for line in db.shipping_lines.find({}, {'name': 1})}
    
    records = df.to_dict('records')
    inserted = 0
    updated = 0
    errors = []

    for index, record in enumerate(records):
        try:
            # Skip empty rows
            if pd.isna(record['name']) or pd.isna(record['contact_email']):
                continue
                
            # Clean data
            cleaned_record = {
                'name': str(record['name']).strip(),
                'contact_email': str(record['contact_email']).strip(),
                'website': str(record.get('website', '')).strip() if not pd.isna(record.get('website')) else '',
                'updated_at': datetime.utcnow()
            }
            
            # Only process new shipping lines
            if cleaned_record['name'] not in existing_names:
                cleaned_record['created_at'] = datetime.utcnow()
                
                result = db.shipping_lines.insert_one(cleaned_record)
                if result.inserted_id:
                    inserted += 1
                    
        except Exception as e:
            errors.append(f"Row {index + 1}: {str(e)}")
            print(f"Error processing record: {record}")
            print(f"Error: {str(e)}")

    return {
        "message": "Bulk upload completed",
        "inserted": inserted,
        "skipped": len(existing_names),
        "errors": errors if errors else None
    }

def bulk_upload_rates(df):
    required_columns = ['shipping_line_id', 'pol_id', 'pod_id', 'container_type', 'rate', 'valid_from', 'valid_to']
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")

    # Get existing rates
    existing_rates = {}
    for rate in db.rates.find({}, {
        'shipping_line_id': 1, 'pol_id': 1, 'pod_id': 1, 
        'container_type': 1, 'rate': 1, 'valid_from': 1, 'valid_to': 1
    }):
        key = f"{rate['shipping_line_id']}-{rate['pol_id']}-{rate['pod_id']}-{rate['container_type']}"
        if key not in existing_rates or rate['valid_to'] > existing_rates[key]['valid_to']:
            existing_rates[key] = rate
    
    records = df.to_dict('records')
    inserted = 0
    updated = 0
    errors = []

    for index, record in enumerate(records):
        try:
            # Skip empty rows
            if any(pd.isna(record[col]) for col in required_columns):
                continue
                
            # Clean data
            cleaned_record = {
                'shipping_line_id': str(record['shipping_line_id']).strip(),
                'pol_id': str(record['pol_id']).strip(),
                'pod_id': str(record['pod_id']).strip(),
                'container_type': str(record['container_type']).strip(),
                'rate': float(record['rate']),
                'valid_from': pd.to_datetime(record['valid_from']).strftime('%Y-%m-%d'),
                'valid_to': pd.to_datetime(record['valid_to']).strftime('%Y-%m-%d'),
                'created_at': datetime.utcnow(),
                'updated_at': datetime.utcnow()
            }
            
            # Generate key for rate
            rate_key = f"{cleaned_record['shipping_line_id']}-{cleaned_record['pol_id']}-{cleaned_record['pod_id']}-{cleaned_record['container_type']}"
            
            if rate_key in existing_rates:
                # Update existing rate by inserting new record
                existing_rate = existing_rates[rate_key]
                if existing_rate['rate'] != cleaned_record['rate']:
                    # Insert new rate record
                    result = db.rates.insert_one(cleaned_record)
                    if result.inserted_id:
                        updated += 1
            else:
                # Insert new rate
                result = db.rates.insert_one(cleaned_record)
                if result.inserted_id:
                    inserted += 1
                    
        except Exception as e:
            errors.append(f"Row {index + 1}: {str(e)}")
            print(f"Error processing record: {record}")
            print(f"Error: {str(e)}")

    return {
        "message": "Bulk upload completed",
        "inserted": inserted,
        "updated": updated,
        "errors": errors if errors else None
    }